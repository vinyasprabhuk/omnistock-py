"""
Port of src/lib/excel/exportWorkbook.ts. Identical sheet names, column
headers/widths, and bold header row to the live app's exports.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_DEPT_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF", name="Arial")
_DEPT_HEADER_FILL = PatternFill("solid", fgColor="2F5233")
_SUBHEADER_FONT = Font(bold=True, name="Arial")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9E4DD")
_TITLE_FONT = Font(bold=True, size=14, name="Arial")
_LABEL_FONT = Font(bold=True, size=11, name="Arial")
_NOTE_FONT = Font(italic=True, size=9, color="808080", name="Arial")


def _sheet(wb: Workbook, name: str, columns: list[tuple[str, int]]):
    sheet = wb.create_sheet(name)
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = width
    return sheet


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_tracker_workbook(date: str, rows: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = _sheet(wb, f"Tracker {date}", [
        ("Item", 24), ("Opening", 12), ("Price", 10), ("Purchase", 12),
        ("Kitchen Requirement", 18), ("Issued", 12), ("Closing", 12),
        ("Unit", 10), ("Usage Cost", 14),
    ])
    for r in rows:
        sheet.append([
            r["itemName"], r["opening"], r["price"], r["purchased"],
            r["kitchenRequirement"], r["issued"], r["closing"], r["unit"], r["usageCost"],
        ])
    return _to_bytes(wb)


def build_master_inventory_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = _sheet(wb, "Master Inventory", [
        ("Item Name", 24), ("Unit", 10), ("Price", 10), ("Opening Stock", 14),
        ("Total Purchased", 16), ("Kitchen Requirement", 18), ("Total Issued", 14),
        ("Current Stock", 14), ("Usage Cost", 14), ("Store Value", 14),
    ])
    for r in rows:
        sheet.append([
            r["itemName"], r["unit"], r["purchasePrice"], r["opening"], r["totalPurchased"],
            r["totalKitchenRequirement"], r["totalIssued"], r["currentStock"], r["usageCost"], r["storeValue"],
        ])
    return _to_bytes(wb)


def build_purchase_order_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = _sheet(wb, "Purchase Order", [
        ("Item Name", 24), ("Unit", 10), ("Current Stock", 14),
        ("Threshold", 14), ("Suggested Order Qty", 18),
    ])
    for r in rows:
        sheet.append([
            r["itemName"], r["unit"], r["currentStock"], r["threshold"], r["orderQty"],
        ])
    return _to_bytes(wb)


def build_kitchen_requirement_workbook(date: str, department_sections: list[dict]) -> bytes:
    """Mirrors the layout of the blank workbook the kitchen team fills in
    and uploads (title, Date: row, note row, then one S.No/Item/Unit/Qty/
    Stock block per department side by side) -- confirmed with the user
    against a real sample of that upload template, so the confirmed-and-
    edited requirement can be handed back to the kitchen team in the same
    format they already recognize."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirement"

    ws["A1"] = "OmniStock — Kitchen Requirement"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Date:"
    ws["A2"].font = _LABEL_FONT
    ws["B2"] = date
    ws["A3"] = "Confirmed kitchen requirement, as saved in OmniStock (Stock is current on-hand, for reference)."
    ws["A3"].font = _NOTE_FONT

    header_row, subheader_row, first_data_row = 5, 6, 7
    block_stride = 6  # 5 data columns (S.No/Item/Unit/Qty/Stock) + 1 spacer

    for i, dept in enumerate(department_sections):
        start_col = i * block_stride + 1
        end_col = start_col + 4

        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = _DEPT_HEADER_FILL
            ws.column_dimensions[cell.column_letter].width = 13
        ws.cell(row=header_row, column=start_col, value=dept["departmentName"]).font = _DEPT_HEADER_FONT
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=end_col)

        for offset, label in enumerate(("S.No", "Item", "Unit", "Qty", "Stock")):
            cell = ws.cell(row=subheader_row, column=start_col + offset, value=label)
            cell.font = _SUBHEADER_FONT
            cell.fill = _SUBHEADER_FILL

        for row_idx, item in enumerate(dept["items"]):
            r = first_data_row + row_idx
            ws.cell(row=r, column=start_col, value=row_idx + 1)
            ws.cell(row=r, column=start_col + 1, value=item["itemName"])
            ws.cell(row=r, column=start_col + 2, value=item["unit"])
            ws.cell(row=r, column=start_col + 3, value=item["qty"])
            ws.cell(row=r, column=start_col + 4, value=item["currentStock"])

    return _to_bytes(wb)


def build_intent_workbook(date: str, dish_counts: list[dict], recipe_prep: list[dict], ingredients: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    prep_sheet = _sheet(wb, "How Much To Prepare", [
        ("Recipe", 28), ("Amount (Litres)", 16), ("Batches Needed", 16), ("Batch Size", 20), ("Driven By", 60),
    ])
    for rp in recipe_prep:
        prep_sheet.append([
            rp["recipeName"], rp["totalLitres"], rp["batchesNeeded"], rp["batchSizeLabel"],
            ", ".join(rp["contributors"]),
        ])

    dish_sheet = _sheet(wb, "Predicted Dish Counts", [
        ("Dish", 32), ("Category", 18), ("Final Qty", 12), ("Source", 10),
    ])
    for d in dish_counts:
        dish_sheet.append([d["dishName"], d["category"], d["finalQty"], d["source"]])

    ing_sheet = _sheet(wb, "Ingredient Requirement", [
        ("Recipe", 24), ("Item", 24), ("Unit", 10), ("Qty", 12), ("Source", 10),
    ])
    for i in ingredients:
        ing_sheet.append([i["groupLabel"], i["itemName"], i["unit"], i["qty"], i["source"]])

    return _to_bytes(wb)
