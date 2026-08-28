"""
Port of src/lib/excel/importWorkbook.ts -- the legacy 31-sheet workbook
importer. Fixed column positions (not pattern-based like the kitchen
parsers): Master Stock sheet A/B/C/D = name/unit/price/opening; day sheets
named "1".."31" use D/E = purchaseQty/usageQty, cross-checked against Master
Stock's row-position link. Only raw inputs are pulled, never formulas.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import TypedDict

from openpyxl import load_workbook

_DAY_SHEET_RE = re.compile(r"^\d+$")


class ImportedItem(TypedDict):
    row: int
    name: str
    unit: str
    purchasePrice: float
    openingStock: float


class ImportedDayTransaction(TypedDict):
    sheetName: str
    date: str
    row: int
    purchaseQty: float
    usageQty: float


class ParsedWorkbook(TypedDict):
    items: list[ImportedItem]
    dayTransactions: list[ImportedDayTransaction]
    daySheetsFound: list[str]
    warnings: list[str]


def _cell_number(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v)


def _cell_date(v):
    if isinstance(v, datetime):
        return v
    if hasattr(v, "isoformat") and not isinstance(v, str):  # date object
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                continue
    return None


def parse_workbook(file_bytes: bytes) -> ParsedWorkbook:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

    warnings: list[str] = []
    items: list[ImportedItem] = []

    if "Master Stock" not in workbook.sheetnames:
        raise ValueError('Sheet "Master Stock" not found in this workbook')
    master_sheet = workbook["Master Stock"]

    for row_number in range(2, (master_sheet.max_row or 0) + 1):
        name = _cell_text(master_sheet.cell(row=row_number, column=1).value)
        if not name:
            continue
        items.append({
            "row": row_number, "name": name,
            "unit": _cell_text(master_sheet.cell(row=row_number, column=2).value) or "UNIT",
            "purchasePrice": _cell_number(master_sheet.cell(row=row_number, column=3).value),
            "openingStock": _cell_number(master_sheet.cell(row=row_number, column=4).value),
        })

    name_by_row = {i["row"]: i["name"] for i in items}
    day_transactions: list[ImportedDayTransaction] = []
    day_sheets_found: list[str] = []

    for sheet_name in workbook.sheetnames:
        if not _DAY_SHEET_RE.match(sheet_name):
            continue  # skips "Master Stock" and any stray non-numeric sheet (e.g. "test")
        day_sheets_found.append(sheet_name)
        sheet = workbook[sheet_name]

        sheet_date = None
        for row_number in range(2, (sheet.max_row or 0) + 1):
            if sheet_date:
                break
            d = _cell_date(sheet.cell(row=row_number, column=1).value)
            if d:
                sheet_date = d

        if not sheet_date:
            warnings.append(f'Day sheet "{sheet_name}" has no date in column A -- skipped')
            continue
        date_key = sheet_date.strftime("%Y-%m-%d")

        for row_number in range(2, (sheet.max_row or 0) + 1):
            purchase_qty = _cell_number(sheet.cell(row=row_number, column=4).value)
            usage_qty = _cell_number(sheet.cell(row=row_number, column=5).value)
            if purchase_qty == 0 and usage_qty == 0:
                continue

            expected_name = name_by_row.get(row_number)
            actual_name = _cell_text(sheet.cell(row=row_number, column=2).value)
            if expected_name and actual_name and expected_name.strip().upper() != actual_name.strip().upper():
                warnings.append(f'Sheet "{sheet_name}" row {row_number}: expected "{expected_name}" but found "{actual_name}" -- skipped')
                continue
            if not expected_name:
                continue

            day_transactions.append({
                "sheetName": sheet_name, "date": date_key, "row": row_number,
                "purchaseQty": purchase_qty, "usageQty": usage_qty,
            })

    day_sheets_found.sort(key=lambda s: int(s))
    return {"items": items, "dayTransactions": day_transactions, "daySheetsFound": day_sheets_found, "warnings": warnings}
