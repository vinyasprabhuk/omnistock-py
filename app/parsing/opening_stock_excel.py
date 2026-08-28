"""
Port of src/lib/parsing/parseOpeningStockExcel.ts. Only the first sheet is
read. This is the one parser in the app that scans header TEXT (ITEM/NAME +
OPEN/STOCK/QTY, first 5 rows only) -- the kitchen-requirement parsers are
pattern-based instead, since real kitchen sheets often lack headers.
"""
from __future__ import annotations

import io
import re
from typing import TypedDict

from openpyxl import load_workbook

_ITEM_HEADER_RE = re.compile(r"ITEM|NAME")
_QTY_HEADER_RE = re.compile(r"OPEN|STOCK|QTY")


class ParsedOpeningStockRow(TypedDict):
    itemText: str
    qty: float


def _cell_text(sheet, row: int, col: int) -> str:
    v = sheet.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _cell_number(sheet, row: int, col: int) -> float | None:
    text = _cell_text(sheet, row, col)
    try:
        return float(text)
    except ValueError:
        return None


def parse_opening_stock_excel(file_bytes: bytes) -> list[ParsedOpeningStockRow]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if not workbook.worksheets:
        return []
    sheet = workbook.worksheets[0]

    item_col, qty_col, header_row = 1, 2, 0

    max_scan_row = min(sheet.max_row or 0, 5)
    found = False
    for r in range(1, max_scan_row + 1):
        if found:
            break
        for c in range(1, (sheet.max_column or 0) + 1):
            text = _cell_text(sheet, r, c).upper()
            if _ITEM_HEADER_RE.search(text):
                for cc in range(1, (sheet.max_column or 0) + 1):
                    qty_text = _cell_text(sheet, r, cc).upper()
                    if _QTY_HEADER_RE.search(qty_text):
                        item_col, qty_col, header_row = c, cc, r
                        found = True
                        break
            if found:
                break

    rows: list[ParsedOpeningStockRow] = []
    for r in range(header_row + 1, (sheet.max_row or 0) + 1):
        item_text = _cell_text(sheet, r, item_col)
        if not item_text:
            continue
        qty = _cell_number(sheet, r, qty_col)
        if qty is None:
            continue
        rows.append({"itemText": item_text, "qty": qty})

    return rows
