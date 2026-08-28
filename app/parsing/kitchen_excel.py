"""
Port of src/lib/parsing/parseKitchenExcel.ts.

Direct structured parse of a kitchen requirement workbook -- no AI. Real
kitchen workbooks don't reliably have "S.NO/ITEM/QTY" text headers, so this
scans every column for the actual repeating pattern that's always true: a
small integer, immediately followed by item text, immediately followed by a
quantity-like value. See the original TS file's comments for the full
rationale -- ported here with identical heuristics and constants.
"""
from __future__ import annotations

import io
import re
from typing import TypedDict

from openpyxl import load_workbook

from app.services.normalize_unit import CANONICAL_UNITS

_QTY_UNIT_RE = re.compile(r"^([\d.]+)\s*([A-Za-z]*)$")
_SMALL_INT_RE = re.compile(r"^\d{1,3}$")
_ALL_DIGITS_RE = re.compile(r"^\d+$")
_QTY_ONLY_RE = re.compile(r"^[\d.]+$")
_SHEET_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})(?:-\d+)?$")

NON_DEPARTMENT_LABELS = {"S.NO", "SNO", "S NO", "SL.NO", "ITEM", "QTY", "ITEMS"}


class ExtractedItem(TypedDict):
    raw_item: str
    quantity: float
    unit: str | None


class ExtractedDepartment(TypedDict):
    name: str
    items: list[ExtractedItem]


class KitchenRequirementExtraction(TypedDict):
    departments: list[ExtractedDepartment]


def _cell_text(sheet, row: int, col: int) -> str:
    if row < 1 or col < 1:
        return ""
    cell = sheet.cell(row=row, column=col)
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        # Match JS's String(number) for whole floats read from Excel (openpyxl
        # gives Python int/float already, unlike ExcelJS's formula-result objects).
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if hasattr(v, "isoformat"):  # datetime
        return v.isoformat()
    return str(v).strip()


def _parse_qty_unit(text: str) -> dict | None:
    match = _QTY_UNIT_RE.match(text.strip())
    if not match:
        return None
    try:
        qty = float(match.group(1))
    except ValueError:
        return None
    if qty != qty or qty <= 0:  # NaN check
        return None
    unit = match.group(2).upper() if match.group(2) else None
    return {"qty": qty, "unit": unit}


def _is_known_unit_text(text: str) -> bool:
    return text.strip().upper() in CANONICAL_UNITS


def _is_small_positive_int(text: str) -> bool:
    if not _SMALL_INT_RE.match(text):
        return False
    n = int(text)
    return 0 <= n <= 500


class _RawBlockRow(TypedDict):
    row: int
    col: int
    itemText: str
    qty: float | None
    unit: str | None


def _find_department_name(sheet, start_row: int, col: int) -> str:
    for back in range(1, 4):
        r = start_row - back
        if r < 1:
            break
        for c in (col, col + 1):
            text = _cell_text(sheet, r, c)
            if text and text.upper() not in NON_DEPARTMENT_LABELS and not _ALL_DIGITS_RE.match(text):
                return text
    return f"Section {col}"


def _extract_block_rows(sheet) -> list[_RawBlockRow]:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    rows: list[_RawBlockRow] = []

    for c in range(1, max_col + 1):
        for r in range(1, max_row + 1):
            num_text = _cell_text(sheet, r, c)
            if not _is_small_positive_int(num_text):
                continue

            item_text = _cell_text(sheet, r, c + 1)
            if not item_text or _ALL_DIGITS_RE.match(item_text):
                continue

            unit_col_text = _cell_text(sheet, r, c + 2)
            qty_only_text = _cell_text(sheet, r, c + 3)
            if _is_known_unit_text(unit_col_text) and _QTY_ONLY_RE.match(qty_only_text):
                try:
                    qty = float(qty_only_text)
                except ValueError:
                    qty = None
                if qty is not None and qty > 0:
                    rows.append({"row": r, "col": c, "itemText": item_text, "qty": qty,
                                 "unit": unit_col_text.strip().upper()})
                    continue

            qty_text = _cell_text(sheet, r, c + 2)
            parsed = _parse_qty_unit(qty_text)
            rows.append({"row": r, "col": c, "itemText": item_text,
                         "qty": parsed["qty"] if parsed else None,
                         "unit": parsed["unit"] if parsed else None})
    return rows


def _group_into_blocks(sheet, rows: list[_RawBlockRow]) -> KitchenRequirementExtraction:
    by_column: dict[int, list[_RawBlockRow]] = {}
    for r in rows:
        by_column.setdefault(r["col"], []).append(r)

    departments_map: dict[str, ExtractedDepartment] = {}

    for col, col_rows in by_column.items():
        col_rows.sort(key=lambda r: r["row"])
        run: list[_RawBlockRow] = []

        def flush():
            nonlocal run
            if not run:
                return
            with_qty = [r for r in run if r["qty"] is not None]
            if with_qty:
                dept_name = _find_department_name(sheet, run[0]["row"], col)
                key = dept_name.strip().upper()
                dept = departments_map.setdefault(key, {"name": dept_name.strip(), "items": []})
                for r in with_qty:
                    dept["items"].append({"raw_item": r["itemText"], "quantity": r["qty"], "unit": r["unit"]})
            run = []

        for r in col_rows:
            if run and r["row"] - run[-1]["row"] > 1:
                flush()
            run.append(r)
        flush()

    return {"departments": list(departments_map.values())}


def _sheet_date_key(sheet_name: str) -> int | None:
    match = _SHEET_DATE_RE.match(sheet_name.strip())
    if not match:
        return None
    d, m, y = match.groups()[:3]
    return int(y) * 10000 + int(m) * 100 + int(d)


def parse_kitchen_excel(file_bytes: bytes) -> KitchenRequirementExtraction:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

    dated = [(sheet, _sheet_date_key(sheet.title)) for sheet in workbook.worksheets]
    dated = [(s, k) for s, k in dated if k is not None]

    if len(dated) >= 2:
        latest_sheet = max(dated, key=lambda sk: sk[1])[0]
        sheets_to_scan = [latest_sheet]
    else:
        sheets_to_scan = workbook.worksheets

    departments_map: dict[str, ExtractedDepartment] = {}
    for sheet in sheets_to_scan:
        rows = _extract_block_rows(sheet)
        extraction = _group_into_blocks(sheet, rows)
        for dept in extraction["departments"]:
            key = dept["name"].strip().upper()
            entry = departments_map.setdefault(key, {"name": dept["name"], "items": []})
            entry["items"].extend(dept["items"])

    return {"departments": list(departments_map.values())}
