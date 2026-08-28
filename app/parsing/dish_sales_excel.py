"""
Parser for the POS "Outlet-Item Wise Report (Row)" export -- both the
multi-month aggregate and the single-day exports share this exact layout:
row 1 = "Date: <range>", row 5 = column headers, data rows from row 10
onward (rows 6-9 are Total/Min/Max/Avg summary rows to skip), with
Sub Total marker rows splitting Taxable/Non-Taxable blocks partway through.
"""
from __future__ import annotations

from datetime import datetime
from typing import TypedDict

from openpyxl import load_workbook


class DishSaleRow(TypedDict):
    date: str  # YYYY-MM-DD, single day
    restaurant: str
    category: str
    item: str
    qty: float


def parse_dish_sales_file(path: str) -> list[DishSaleRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    date_cell = ws["B1"].value or ""
    date_from = str(date_cell).split(" to ")[0].strip()
    datetime.strptime(date_from, "%Y-%m-%d")  # raises if the export format changed

    rows: list[DishSaleRow] = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        marker, restaurant, category, item, qty = row[0], row[1], row[2], row[3], row[4]
        if marker in ("Sub Total", "Total", "Min.", "Max.", "Avg."):
            continue
        if not item or qty is None:
            continue
        rows.append({
            "date": date_from, "restaurant": str(restaurant or "").strip(),
            "category": str(category or "").strip(), "item": str(item).strip(),
            "qty": float(qty),
        })
    wb.close()
    return rows
