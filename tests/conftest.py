"""
Shared pytest fixtures.

Every fixture that touches a database works against a throwaway COPY of the
real (read-only) reference database at instance/reference/dev.db.pristine --
never the live working copy at instance/dev.db. Tests are free to write,
mutate, and leave a mess; each test gets its own fresh copy via pytest's
built-in `tmp_path` fixture, so nothing persists between tests and nothing
ever touches real data.

`db_conn` and `client` (and the `app` it comes from) all resolve to the SAME
underlying database FILE for a given test, via the shared `db_path` fixture
below -- otherwise a test that creates a row via `db_conn` and then expects
to see it through an HTTP request via `client` would silently be looking at
two different files.
"""
from __future__ import annotations

import shutil
import sqlite3
from pathlib import Path

import pytest

PRISTINE_DB = Path(__file__).resolve().parent.parent / "instance" / "reference" / "dev.db.pristine"


@pytest.fixture()
def db_path(tmp_path) -> Path:
    if not PRISTINE_DB.exists():
        pytest.skip(f"reference DB not found at {PRISTINE_DB} -- run tools/verify_schema.py setup first")
    path = tmp_path / "test.db"
    shutil.copy(PRISTINE_DB, path)
    # PRISTINE_DB is deliberately chmod 444 (see tools/ setup) so it can never
    # be written to by accident; shutil.copy preserves that on most systems,
    # so the copy needs to be made writable explicitly.
    path.chmod(0o644)
    return path


@pytest.fixture()
def db_conn(db_path):
    """A live sqlite3 connection to the test's shared throwaway DB copy."""
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    yield conn
    conn.close()


@pytest.fixture()
def branch_id(db_conn) -> str:
    return db_conn.execute("SELECT id FROM Branch LIMIT 1").fetchone()[0]


@pytest.fixture()
def admin_user_id(db_conn) -> str:
    return db_conn.execute("SELECT id FROM User WHERE role = 'ADMIN' LIMIT 1").fetchone()[0]


@pytest.fixture()
def app(db_path, monkeypatch):
    """A Flask app instance pointed at this test's shared throwaway DB file."""
    monkeypatch.setenv("DATABASE_PATH", str(db_path))

    from app import create_app
    flask_app = create_app()
    flask_app.config.update(TESTING=True)
    return flask_app


@pytest.fixture()
def client(app):
    return app.test_client()


def login(client, username: str, password: str):
    """Log the test client in, following the real GET-token-then-POST flow
    (not skipping CSRF) so tests exercise the actual auth path."""
    import re
    resp = client.get("/login")
    match = re.search(r'name="_csrf_token" value="([^"]*)"', resp.get_data(as_text=True))
    token = match.group(1)
    return client.post("/login", data={"username": username, "password": password, "_csrf_token": token})


def csrf_token(client) -> str:
    """Fetch a fresh CSRF token for a POST from an already-logged-in client
    (mirrors what the JS on real pages does before a long-lived form
    submits) -- avoids scraping it out of a specific page's HTML, which
    breaks if that page's markup changes."""
    return client.get("/api/csrf-token").get_json()["csrfToken"]


# --- Full current-schema fixtures --------------------------------------
# The pristine reference DB predates the Intent/Recipe tables and the
# KitchenRequirement reject-with-comment columns (it was captured for
# golden-parity numeric checks against the original Next.js app, not for
# full-app e2e coverage) -- db_path/db_conn/app/client above deliberately
# use it AS-IS so those parity tests keep testing exactly what they were
# captured against. For everything else, full_db_path applies every
# guarded/idempotent migration on top of a fresh pristine copy, so tests
# see the exact schema production runs today.
MIGRATIONS = (
    "migrate_add_intent_recipe",
    "migrate_add_workstation",
    "migrate_add_requirement_review",
    "migrate_kitchen_requirement_v2",
)


@pytest.fixture()
def full_db_path(db_path) -> Path:
    import importlib
    for name in MIGRATIONS:
        mod = importlib.import_module(f"tools.{name}")
        mod.migrate(db_path)
    return db_path


@pytest.fixture()
def full_db_conn(full_db_path):
    conn = sqlite3.connect(str(full_db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    yield conn
    conn.close()


@pytest.fixture()
def full_app(full_db_path, monkeypatch):
    monkeypatch.setenv("DATABASE_PATH", str(full_db_path))
    from app import create_app
    flask_app = create_app()
    flask_app.config.update(TESTING=True)
    return flask_app


@pytest.fixture()
def full_client(full_app):
    return full_app.test_client()


def build_kitchen_upload_xlsx(department_items: dict[str, list[tuple[str, float, str]]]) -> bytes:
    """Builds a real .xlsx in the exact block layout parse_kitchen_excel
    expects (a small integer, then item text, then unit, then qty, each
    one cell over -- matched by scanning every column for that repeating
    pattern, not by header text). department_items: {dept_name: [(item_
    name, qty, unit), ...]}. Real item names (matching real Item Master
    rows) let the AUTO match path exercise for real, same as a genuine
    upload."""
    import io
    from datetime import datetime
    from openpyxl import Workbook

    wb = Workbook()
    # openpyxl stamps wall-clock created/modified timestamps into every
    # workbook by default, so two calls with IDENTICAL cell content would
    # otherwise still produce different file bytes (and therefore
    # different sha256 hashes) if they land in different seconds --
    # fixing both to a constant makes identical content always hash
    # identically, which duplicate-upload-detection tests depend on.
    wb.properties.created = datetime(2026, 1, 1)
    wb.properties.modified = datetime(2026, 1, 1)
    ws = wb.active
    col = 1
    for dept_name, items in department_items.items():
        ws.cell(row=5, column=col, value=dept_name)
        ws.cell(row=6, column=col, value="S.No")
        ws.cell(row=6, column=col + 1, value="Item")
        ws.cell(row=6, column=col + 2, value="Unit")
        ws.cell(row=6, column=col + 3, value="Qty")
        for i, (name, qty, unit) in enumerate(items):
            r = 7 + i
            ws.cell(row=r, column=col, value=i + 1)
            ws.cell(row=r, column=col + 1, value=name)
            ws.cell(row=r, column=col + 2, value=unit)
            ws.cell(row=r, column=col + 3, value=qty)
        col += 6
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_user(conn, role: str, branch_id: str | None, department_id: str | None = None) -> tuple[str, str, str]:
    """Creates a real, correctly-hashed test user of the given role.
    Returns (user_id, username, password). Every call gets a unique
    username so tests can create as many as they need without collisions."""
    import uuid
    from app.security import hash_password
    user_id = uuid.uuid4().hex
    username = f"test_{role.lower()}_{uuid.uuid4().hex[:8]}"
    password = "testpass123"
    conn.execute(
        "INSERT INTO User (id, name, email, passwordHash, role, branchId, departmentId, active, createdAt, updatedAt) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 1, datetime('now'), datetime('now'))",
        (user_id, f"Test {role}", username, hash_password(password), role, branch_id, department_id),
    )
    conn.commit()
    return user_id, username, password
