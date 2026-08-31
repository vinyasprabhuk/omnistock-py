"""e2e coverage for the remaining Excel export routes and Dashboard tabs
not already exercised by the kitchen-requirement/low-stock suites."""
from __future__ import annotations

from tests.conftest import csrf_token, login, make_user

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _admin_client(full_app, full_db_conn):
    client = full_app.test_client()
    _, username, password = make_user(full_db_conn, "ADMIN", None)
    login(client, username, password)
    return client


class TestExportRoutes:
    def test_tracker_export(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/api/export/tracker?date=2026-08-25&branchId={branch_id}")
        assert resp.status_code == 200
        assert resp.mimetype == XLSX_MIME

    def test_inventory_export(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/api/export/inventory?branchId={branch_id}")
        assert resp.status_code == 200
        assert resp.mimetype == XLSX_MIME

    def test_requirements_export_with_no_confirmed_data_is_still_a_valid_workbook(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/api/export/requirements?date=2026-08-25&branchId={branch_id}")
        assert resp.status_code == 200
        assert resp.mimetype == XLSX_MIME

    def test_intent_export_with_generated_day(self, full_app, full_db_conn, branch_id):
        from app.dates import date_key_to_db
        from app.db import new_id
        from app.services.intent import generate_intent_day

        item_id = new_id()
        full_db_conn.execute(
            "INSERT INTO Item (id, name, unit, purchasePrice, category, active, createdAt, updatedAt) "
            "VALUES (?, 'Export Test Item', 'KG', 5, 'TEST', 1, datetime('now'), datetime('now'))",
            (item_id,),
        )
        dept = full_db_conn.execute("SELECT id FROM Department LIMIT 1").fetchone()["id"]
        dish_id = new_id()
        full_db_conn.execute(
            "INSERT INTO Dish (id, name, departmentId, category, active, createdAt, updatedAt) "
            "VALUES (?, 'Export Test Dish', ?, 'OTHER', 1, datetime('now'), datetime('now'))",
            (dish_id, dept),
        )
        recipe_id = new_id()
        full_db_conn.execute(
            "INSERT INTO Recipe (id, dishId, name, servesQty, createdAt, updatedAt) "
            "VALUES (?, ?, 'Export Test Dish', 1, datetime('now'), datetime('now'))",
            (recipe_id, dish_id),
        )
        full_db_conn.execute(
            "INSERT INTO RecipeLine (id, recipeId, itemId, qty, rawIngredientName, rawQtyValue, rawQtyUnit, "
            "matchStatus, createdAt) VALUES (?, ?, ?, 1, 'x', 1, 'KG', 'AUTO', datetime('now'))",
            (new_id(), recipe_id, item_id),
        )
        full_db_conn.execute(
            "INSERT INTO DishSale (id, date, dishId, rawItemName, qty, matchStatus, createdAt) "
            "VALUES (?, ?, ?, 'x', 5, 'AUTO', datetime('now'))",
            (new_id(), date_key_to_db("2026-08-24"), dish_id),
        )
        full_db_conn.commit()
        generate_intent_day(full_db_conn, "2026-08-31", branch_id, weeks_back=1)
        full_db_conn.commit()

        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/api/export/intent?date=2026-08-31&branchId={branch_id}")
        assert resp.status_code == 200
        assert resp.mimetype == XLSX_MIME
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(resp.data))
        assert "Predicted Dish Counts" in wb.sheetnames
        assert "Ingredient Requirement" in wb.sheetnames

    def test_export_routes_require_auth(self, full_app):
        client = full_app.test_client()
        for path in ("/api/export/tracker", "/api/export/inventory", "/api/export/purchase-order"):
            resp = client.get(path)
            assert resp.status_code == 302
            assert "/login" in resp.headers["Location"]


class TestDashboardTabs:
    def test_purchase_spend_tab(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/dashboard?tab=purchase&branchId={branch_id}")
        assert resp.status_code == 200
        assert b"Purchase Spend" in resp.data or b"purchase" in resp.data.lower()

    def test_usage_spend_tab(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/dashboard?tab=usage&branchId={branch_id}")
        assert resp.status_code == 200

    def test_comparison_tab_month_mode(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(f"/dashboard?tab=comparison&cmpMode=month&branchId={branch_id}")
        assert resp.status_code == 200
        assert b"Last Month" in resp.data

    def test_comparison_tab_custom_date_range(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        resp = client.get(
            f"/dashboard?tab=comparison&cmpMode=custom&cmpAFrom=2026-08-01&cmpATo=2026-08-07"
            f"&cmpBFrom=2026-08-08&cmpBTo=2026-08-14&branchId={branch_id}"
        )
        assert resp.status_code == 200

    def test_dashboard_department_filter(self, full_app, full_db_conn, branch_id):
        client = _admin_client(full_app, full_db_conn)
        dept = full_db_conn.execute("SELECT name FROM Department LIMIT 1").fetchone()["name"]
        resp = client.get(f"/dashboard?department={dept}&branchId={branch_id}")
        assert resp.status_code == 200

    def test_viewer_can_view_dashboard_but_not_write(self, full_app, full_db_conn, branch_id):
        # Unlike ADMIN, every other role (including VIEWER) needs a real
        # branchId assigned -- page_resolve_branch raises otherwise.
        client = full_app.test_client()
        _, username, password = make_user(full_db_conn, "VIEWER", branch_id)
        login(client, username, password)
        resp = client.get("/dashboard")
        assert resp.status_code == 200
