"""Correctness tests for the low-stock threshold calculation
(app.services.calculations.get_low_stock): threshold = 30% of opening
stock, flagged when currentStock <= threshold. Exact numeric
assertions, not just "the page loads" -- this is the formula the
Dashboard's Low Stock tab and the Purchase Order export both depend on,
and it's already been the source of one real ambiguity this project hit
("30% of actual stock" vs "30% of opening stock") -- worth locking down
precisely.
"""
from __future__ import annotations

from app.dates import date_key_to_db
from app.services.calculations import get_low_stock


_item_counter = [0]


def _new_item(db_conn):
    """A brand-new synthetic item with zero pre-existing purchase/issue
    history -- reusing a real item from the pristine snapshot would mean
    its already-baked-in Purchase/StockIssue rows silently change the
    expected currentStock out from under the test."""
    _item_counter[0] += 1
    item_id = f"test-item-{_item_counter[0]}"
    db_conn.execute(
        "INSERT INTO Item (id, name, unit, purchasePrice, category, active, createdAt, updatedAt) "
        "VALUES (?, ?, 'KG', 10, 'TEST', 1, datetime('now'), datetime('now'))",
        (item_id, f"Test Item {_item_counter[0]}"),
    )
    db_conn.commit()
    return {"id": item_id, "unit": "KG"}


def _set_opening(db_conn, item_id, branch_id, qty):
    db_conn.execute("DELETE FROM ItemOpeningStock WHERE itemId = ? AND branchId = ?", (item_id, branch_id))
    db_conn.execute(
        "INSERT INTO ItemOpeningStock (id, itemId, branchId, qty, updatedAt) VALUES (?, ?, ?, ?, datetime('now'))",
        (f"ios-{item_id}", item_id, branch_id, qty),
    )
    db_conn.commit()


def _purchase(db_conn, item_id, branch_id, qty, date_db):
    pid = f"p-{item_id}-{qty}"
    db_conn.execute(
        "INSERT INTO Purchase (id, date, branchId, createdAt) VALUES (?, ?, ?, datetime('now'))",
        (pid, date_db, branch_id),
    )
    db_conn.execute(
        "INSERT INTO PurchaseItem (id, purchaseId, itemId, qty, rate) VALUES (?, ?, ?, ?, 1)",
        (f"pi-{pid}", pid, item_id, qty),
    )
    db_conn.commit()


def _issue(db_conn, item_id, branch_id, qty, date_db):
    dept = db_conn.execute("SELECT id FROM Department LIMIT 1").fetchone()["id"]
    sid = f"s-{item_id}-{qty}"
    db_conn.execute(
        "INSERT INTO StockIssue (id, date, branchId, departmentId, createdAt) VALUES (?, ?, ?, ?, datetime('now'))",
        (sid, date_db, branch_id, dept),
    )
    db_conn.execute(
        "INSERT INTO StockIssueItem (id, stockIssueId, itemId, qty) VALUES (?, ?, ?, ?)",
        (f"si-{sid}", sid, item_id, qty),
    )
    db_conn.commit()


class TestThresholdFormula:
    def test_threshold_is_exactly_30_percent_of_opening(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 100.0)
        rows = get_low_stock(db_conn, branch_id)
        row = next((r for r in rows if r["itemId"] == item["id"]), None)
        # currentStock == opening (100) with no purchase/issue -- well above
        # threshold (30), so should NOT be flagged at all yet.
        assert row is None

    def test_flags_exactly_at_the_boundary(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 100.0)
        # currentStock must land EXACTLY on the 30.0 threshold: issue 70 of
        # the 100 opening, leaving exactly 30 -- the "<=" comparison means
        # this boundary case must flag, not just anything strictly below it.
        _issue(db_conn, item["id"], branch_id, 70.0, date_key_to_db("2026-08-25"))
        rows = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        row = next(r for r in rows if r["itemId"] == item["id"])
        assert row["threshold"] == 30.0
        assert row["currentStock"] == 30.0

    def test_does_not_flag_just_above_threshold(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 100.0)
        _issue(db_conn, item["id"], branch_id, 69.0, date_key_to_db("2026-08-25"))  # leaves 31, above 30
        rows = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        assert not any(r["itemId"] == item["id"] for r in rows)

    def test_threshold_rounds_to_two_decimals(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 10.0)  # 30% of 10 = 3.0 exactly, use an odd number instead
        _set_opening(db_conn, item["id"], branch_id, 7.0)  # 30% of 7 = 2.1 exactly
        rows = get_low_stock(db_conn, branch_id)
        row = next((r for r in rows if r["itemId"] == item["id"]), None)  # currentStock == opening == 7, above 2.1
        assert row is None

        _issue(db_conn, item["id"], branch_id, 4.9, date_key_to_db("2026-08-25"))  # leaves 2.1 exactly
        rows2 = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        row2 = next(r for r in rows2 if r["itemId"] == item["id"])
        assert row2["threshold"] == 2.1

    def test_zero_opening_stock_gives_zero_threshold_flags_only_at_or_below_zero(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 0.0)
        rows = get_low_stock(db_conn, branch_id)
        row = next(r for r in rows if r["itemId"] == item["id"])
        assert row["threshold"] == 0.0
        assert row["currentStock"] == 0.0  # 0 opening, no activity -- flagged since 0 <= 0

    def test_purchases_raise_current_stock_above_threshold(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 10.0)  # threshold 3.0
        _issue(db_conn, item["id"], branch_id, 9.0, date_key_to_db("2026-08-25"))  # leaves 1, below threshold

        flagged_before_purchase = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        assert any(r["itemId"] == item["id"] for r in flagged_before_purchase)

        _purchase(db_conn, item["id"], branch_id, 5.0, date_key_to_db("2026-08-25"))  # now 1+5=6, above 3.0
        flagged_after_purchase = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        assert not any(r["itemId"] == item["id"] for r in flagged_after_purchase)

    def test_negative_current_stock_is_flagged(self, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 5.0)
        _issue(db_conn, item["id"], branch_id, 8.0, date_key_to_db("2026-08-25"))  # -3, below threshold of 1.5
        rows = get_low_stock(db_conn, branch_id, date_key_to_db("2026-08-25"))
        row = next(r for r in rows if r["itemId"] == item["id"])
        assert row["currentStock"] == -3.0


class TestLowStockEndToEnd:
    def test_dashboard_low_stock_tab_shows_correct_count_and_export_link(self, client, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 10.0)
        _issue(db_conn, item["id"], branch_id, 9.0, date_key_to_db("2026-08-25"))

        from tests.conftest import login, make_user
        _, username, password = make_user(db_conn, "ADMIN", None)
        login(client, username, password)

        resp = client.get(f"/dashboard?tab=lowstock&branchId={branch_id}&to=2026-08-25")
        assert resp.status_code == 200
        body = resp.get_data(as_text=True)
        assert "Generate Purchase Order (Excel)" in body

    def test_purchase_order_export_suggests_qty_back_to_opening_stock(self, client, db_conn, branch_id):
        item = _new_item(db_conn)
        _set_opening(db_conn, item["id"], branch_id, 20.0)
        _issue(db_conn, item["id"], branch_id, 19.0, date_key_to_db("2026-08-31"))  # leaves 1, threshold 6

        from tests.conftest import login, make_user
        _, username, password = make_user(db_conn, "ADMIN", None)
        login(client, username, password)

        resp = client.get(f"/api/export/purchase-order?branchId={branch_id}")
        assert resp.status_code == 200
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Purchase Order"]
        target_row = next(
            (row for row in ws.iter_rows(min_row=2, values_only=True) if row[2] == 1.0 and row[3] == 6.0),
            None,
        )
        assert target_row is not None, "expected a row with currentStock=1.0, threshold=6.0"
        assert target_row[4] == 19.0, "suggested order qty should bring stock back to the 20.0 opening baseline"
